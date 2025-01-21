from os.path import splitext, basename, dirname, join, exists, isfile
from os import makedirs, path
from pandas import read_csv, DataFrame, concat, read_excel
import pandas as pd
from numpy import zeros, min as np_min, max as np_max, array, arange, meshgrid, vstack, histogram, finfo, log, argmax, asarray, mean, median
from matplotlib.pyplot import subplots, show, close, savefig, draw, imsave
from matplotlib.widgets import PolygonSelector
from matplotlib.patches import Polygon
from matplotlib.colors import to_rgba
import matplotlib.patches as patches
from PIL import Image, ImageOps
from skimage.filters import threshold_otsu, threshold_yen, threshold_li, threshold_isodata, threshold_mean, threshold_minimum
from skimage.measure import label, regionprops
from skimage.morphology import remove_small_objects
from matplotlib.path import Path
from tqdm import tqdm
from scipy.fftpack import fft2, ifft2, fftshift
import numpy as np
import matplotlib.pyplot as plt
from czifile import CziFile
from aicspylibczi import CziFile as aicCzi
import cv2
import xml.etree.ElementTree as ET
import time
import pickle
from readlif.reader import LifFile
import random

import hashlib
import matplotlib.gridspec as gridspec

import json
import os
import tempfile
import os
import pandas as pd
from tkinter import messagebox, Toplevel, simpledialog, messagebox, ttk, StringVar
from os.path import splitext, basename, dirname, join
from matplotlib.pyplot import Figure
from tkinter.ttk import Button, Label, Entry, OptionMenu, Style, Checkbutton, Frame, Progressbar, Combobox


# Adding current directory to sys.path
import sys, os
current_dir = os.path.dirname(os.path.abspath(__file__))
icon_path = os.path.join(current_dir, "images", "synaptocatcher.ico")
sys.path.append(current_dir)
from graphical_processor import PolygonDrawer, ParallelogramEditor, draw_polygons_on_image, ColorCycler, simplify_contour
from graphical_processor import invert_image, create_region_mask, PolygonModifier, ThemeManager, initialize_window

def generate_hashes_from_metadata(metadata, priority_keys):
    """
    Generates individual hashes for each priority key.

    Args:
        metadata (dict): The metadata dictionary.
        priority_keys (list): List of priority keys for the current step.

    Returns:
        list: List of short hashes corresponding to each priority key.
    """
    hashes = []
    for key in sorted(priority_keys):
        if key in metadata:
            value = metadata[key]
            # Serialize the value to a JSON string for consistent hashing
            value_str = json.dumps(value, sort_keys=True)
            # Create a unique string for hashing per key
            hash_input = f"{key}:{value_str}".encode('utf-8')
            # Generate a short hash (8 characters)
            short_hash = hashlib.sha1(hash_input).hexdigest()[:8]
            hashes.append(short_hash)
        else:
            raise KeyError(f"Priority key '{key}' not found in metadata.")
    return hashes

def extract_hashes_from_filename(filename):
    """
    Extracts the base filename and list of hashes from a given filename.

    Args:
        filename (str): The filename to parse.

    Returns:
        tuple: (base_filename, list_of_hashes)
    """
    base_and_hashes, ext = os.path.splitext(filename)
    if '.' in base_and_hashes:
        base_filename, hashes_str = base_and_hashes.rsplit('.', 1)
        hashes = hashes_str.split('-')
    else:
        base_filename = base_and_hashes
        hashes = []
    return base_filename, hashes

def find_closest_match(hashes, candidates, base_filename):
    """
    Finds the file with the highest number of matching hashes.

    Args:
        hashes (list): List of current priority hashes.
        candidates (list): List of candidate filenames in the directory.
        base_filename (str): The base filename to match.

    Returns:
        str or None: The best matching filename or None if no match found.
    """
    max_common_hashes = -1
    best_candidate = None
    for candidate in candidates:
        candidate_base, candidate_hashes = extract_hashes_from_filename(candidate)
        if candidate_base != base_filename:
            continue
        common_hashes = set(hashes) & set(candidate_hashes)
        if len(common_hashes) > max_common_hashes:
            max_common_hashes = len(common_hashes)
            best_candidate = candidate
    return best_candidate

def hash_convert_path(path, hashes):

    base, ext = os.path.splitext(path)
    directory = os.path.dirname(path) or '.'
    base_filename = os.path.basename(base)
    
    hashes_str = "-".join(hashes)
    filename = f"{base_filename}.{hashes_str}{ext}"
    converted_path = os.path.join(directory, filename)
    return converted_path

def save_image(image, path, Step = None, priority_keys=None, overwrite_similar=False, **kwargs):
    """
    Saves an image with a filename that includes hashes of priority metadata keys.

    Args:
        image: The image to save (PIL Image, NumPy ndarray, or Matplotlib Figure).
        path (str): The base path where the image will be saved.
        priority_keys (list, optional): List of priority metadata keys for hashing.
        overwrite_similar (bool, optional): If True, overwrite the most similar existing image.
        **kwargs: Additional keyword arguments for image saving functions.

    Raises:
        KeyError: If a priority key is missing in the metadata.
        TypeError: If the image type is unsupported.
    """
    metadata = load_metadata()
    base, ext = os.path.splitext(path)
    directory = os.path.dirname(path) or '.'
    base_filename = os.path.basename(base)

    if priority_keys:
        # Generate individual hashes based on priority keys
        hashes = generate_hashes_from_metadata(metadata, priority_keys)
        # Construct new filename with hashes separated by '-'
        hashes_str = "-".join(hashes)
        filename = f"{base_filename}.{hashes_str}{ext}"
    else:
        # No priority keys, use base filename
        filename = f"{base_filename}{ext}"
        hashes = []

    full_path = os.path.join(directory, filename)

    if overwrite_similar and hashes:
        # Search for existing files with the same base filename and any hashes
        existing_files = [
            f for f in os.listdir(directory)
            if f.startswith(f"{base_filename}.") and f.endswith(ext)
        ]
        if existing_files:
            # Find the most similar file based on matching hashes
            best_match = find_closest_match(hashes, existing_files, base_filename)
            if best_match:
                best_match_path = os.path.join(directory, best_match)
                os.remove(best_match_path)
                print(f"Deleted similar image file: {best_match_path}")

    # Save the image
    if isinstance(image, Image.Image):
        image.save(full_path, **kwargs)
    elif isinstance(image, np.ndarray):
        cv2.imwrite(full_path, image)
    elif isinstance(image, Figure):
        image.savefig(full_path)
    else:
        raise TypeError("Unsupported image type. The image must be a PIL Image, a numpy ndarray, or a Matplotlib Figure.")

    # Write metadata to file if available
    if priority_keys:
        selected_metadata = {key: metadata[key] for key in priority_keys if key in metadata}
        
        # Извлекаем 'selected_location' и переименовываем его в 'region' - кастыль
        region_value = selected_metadata.get('selected_location', None)  # Получаем значение или None, если ключ отсутствует
        selected_metadata.pop('selected_location', None)
        
        # Создаем новый словарь с добавлением 'Step' и 'region' в начале
        selected_metadata = {
            'Step': Step,
            'region': region_value,  # Переименованный ключ
            **selected_metadata  # Добавляем оставшиеся ключи
        }
        
        write_data_to_file(full_path, selected_metadata)



def read_image(path, priority_keys=None, as_pil=False, **kwargs):
    """
    Reads an image based on the priority metadata keys.

    Args:
        path (str): The base path to read the image from.
        priority_keys (list, optional): List of priority metadata keys for hashing.
        as_pil (bool, optional): If True, returns a PIL Image; otherwise, returns a NumPy array.
        **kwargs: Additional keyword arguments for image reading functions.

    Returns:
        The loaded image as a PIL Image or a NumPy array.

    Raises:
        FileNotFoundError: If no matching image file is found.
        KeyError: If a priority key is missing in the metadata.
    """
    metadata = load_metadata()
    base, ext = os.path.splitext(path)
    directory = os.path.dirname(path) or '.'
    base_filename = os.path.basename(base)

    if priority_keys:
        # Generate individual hashes based on priority keys
        hashes = generate_hashes_from_metadata(metadata, priority_keys)
        hashes_str = "-".join(hashes)
        expected_filename = f"{base_filename}.{hashes_str}{ext}"
        expected_path = os.path.join(directory, expected_filename)
        print(expected_path)
        if os.path.exists(expected_path):
            path_to_load = expected_path
        else:
            # Search for the closest matching file based on hashes
            existing_files = [
                f for f in os.listdir(directory)
                if f.startswith(f"{base_filename}.") and f.endswith(ext)
            ]
            best_match = find_closest_match(hashes, existing_files, base_filename)
            if best_match:
                path_to_load = os.path.join(directory, best_match)
                print(f"Loaded closest matching image file: {best_match}")
            else:
                print('--------------------')
                print("ERROR:")
                print("No corresponding binarization file found. Try to Filter and Binarize first.")     
                print('--------------------')           
                raise FileNotFoundError("No corresponding binarization file found. Try to Filter and Binarize first.")
    else:
        # No priority keys, use base filename
        path_to_load = os.path.join(directory, f"{base_filename}{ext}")

    # Read the image
    if as_pil:
        return Image.open(path_to_load)
    else:
        return cv2.imread(path_to_load)
    
def extract_scaling_distances_from_czi(filename):
    # Open the .czi file
    czi = aicCzi(filename)
    
    # Get the metadata
    metadata_xml = czi.meta
    metadata_str = ET.tostring(metadata_xml, encoding='utf-8', method='xml')
    
    # Parsing the metadata string
    root = ET.fromstring(metadata_str)
    
    distances = {"X": None, "Y": None, "Z": None}
    
    # Looking for Scaling - Items - Distance
    for scaling in root.findall(".//Scaling/Items/Distance"):
        if 'Id' in scaling.attrib:
            id_value = scaling.attrib['Id']
            value = scaling.find('Value').text
            if id_value == 'X':
                distances["X"] = float(value)
            elif id_value == 'Y':
                distances["Y"] = float(value)
            elif id_value == 'Z':
                distances["Z"] = float(value)
    
    return distances

def filetype_checking(file_path):
    file_extension = splitext(file_path)[1].lower()  
    if file_extension == '.czi':
        n_of_images = 1
    elif file_extension == '.lif':        
        lif = LifFile(file_path)        
        img_list = [i for i in lif.get_iter_image()]    
        n_of_images = len(img_list)
    elif file_extension == '.tif':  
        n_of_images = 1
    return n_of_images

def read_tiff_channels(file_path, num_channels, num_layers):
    # Open the TIFF file
    tiff_image = Image.open(file_path)

    # Check the number of frames (images) in the file
    n_frames = tiff_image.n_frames
    if n_frames != num_channels * num_layers:
        raise ValueError(f"Количество кадров в файле ({n_frames}) не соответствует заданным каналам ({num_channels}) и слоям ({num_layers})")

    # Initialize lists for each channel
    channels = [[] for _ in range(num_channels)]

    # Read all images and distribute them across channels
    for i in range(n_frames):
        tiff_image.seek(i)
        frame = np.array(tiff_image.copy())
        channel_index = i % num_channels
        channels[channel_index].append(frame)

    # Convert lists to numpy arrays for ease of use
    channels = [np.array(channel) for channel in channels]

    # Close the file
    tiff_image.close()

    return channels

def collect_lif_frames(im_in, ch):
    
    z_list = [i for i in im_in.get_iter_z(t=0, c=0)]
    z_n = len(z_list)
        
    channel_list = [i for i in im_in.get_iter_c(t=0, z=0)]
    ch_n = len(channel_list)
    
    def remove_shift(lst, ch):
        shift = ((ch_n-1) * ch - 1) % len(lst)
        reverse_shift = len(lst) - shift
        return lst[reverse_shift:] + lst[:reverse_shift]
    
    ch = (ch_n - ch)
    
    frames_out = []
    for z_real in list(range(z_n)):    
        z=(z_real*ch_n)%z_n       
        c=(z%ch_n+ch)%ch_n    
        frames_out.append(np.array(im_in.get_frame(z = z, c = c)))
    frames_out = np.array(remove_shift(frames_out, ch))
    
    return frames_out, ch_n, z_n

def handle_slide_indexing(slice_start, slice_end, max_slice):
    
    # Handle 'all' for slice_start and slice_end
    if slice_start == 'all' or slice_end == 'all':
        slice_start = 0
        slice_end = max_slice - 1
        
        save_params({'slice_start': slice_start+1, 'slice_end': slice_end+1})
    else:
        slice_start = int(slice_start)
        slice_end = int(slice_end)

    if slice_start < 0 or slice_end >= max_slice:
        print(f"Invalid slice range: start={slice_start}, end={slice_end}. Must be within 0 and {max_slice - 1}.")
        return None
    
    slide = list(range(slice_start, slice_end + 1))
    return slide

def extract_czi_image_stack(file_path, slice_start, slice_end, target_ch, dapi_ch):
    base_name = splitext(basename(file_path))[0]
    experiment_date = basename(dirname(file_path))    
    file_extension = splitext(file_path)[1].lower() 
    
    if file_extension != '.czi':
        print(f"Unsupported file format: {file_extension}")
        return None

    # Convert slice_start and slice_end to numeric values if necessary
    with CziFile(file_path) as czi:
        image_data = czi.asarray()
        im_index = 0
        # Handle image dimension reduction for tiled images (if applicable)
        if len(image_data.shape) == 8:  # Case with tiles
            image_data = image_data[:, :, 0, :, :, :, :, :]
        
        max_channels = image_data.shape[2]
        max_slice = image_data.shape[3]
        
        slide = handle_slide_indexing(slice_start, slice_end, max_slice)
        
        if target_ch >= max_channels:
            print(f"Invalid target channel: {target_ch}. Max channels available: {max_channels}.")
            return None
        
        # Output directory for results
        output_path = join(dirname(file_path), f"{base_name}_results")        
        makedirs(output_path, exist_ok=True)
        
        # Prepare target channel slice
        sample_slice_1 = np.max(image_data[0, 0, target_ch, slide, :, :, 0], axis=0)
        synaptotag_file_path = join(output_path, f"{base_name}_{im_index}_synaptotag.png")
        save_image(sample_slice_1, synaptotag_file_path, Step = "Target data", priority_keys=priority_keys)

        combined_image = np.zeros((*sample_slice_1.shape, 3), dtype='uint8')
        sample_slice_1_normalized = (sample_slice_1 - np.min(sample_slice_1)) / (np.max(sample_slice_1) - np.min(sample_slice_1)) * 255
        combined_image[:, :, 0] = sample_slice_1_normalized.astype('uint8')  # RED: target_ch (synaptotagmin)

        # Check if DAPI channel is valid and add it if applicable
        if dapi_ch < max_channels:
            sample_slice_3 = np.max(image_data[0, 0, dapi_ch, slide, :, :, 0], axis=0)
            sample_slice_3_normalized = (sample_slice_3 - np.min(sample_slice_3)) / (np.max(sample_slice_3) - np.min(sample_slice_3)) * 255
            combined_image[:, :, 2] = sample_slice_3_normalized.astype('uint8')  # BLUE: dapi_ch (cell-label)

        return [combined_image]
    
def extract_lif_stack(file_path, slice_start, slice_end, target_ch, dapi_ch):
    base_name = splitext(basename(file_path))[0]
    experiment_date = basename(dirname(file_path))    
    file_extension = splitext(file_path)[1].lower() 
    
    if file_extension == '.lif':        
        lif = LifFile(file_path)        
        img_list = [i for i in lif.get_iter_image()]
        
        combined_image_s = []
        for im_index, image in enumerate(img_list):      
                        
            frames_1, max_channels, max_slice = collect_lif_frames(image, target_ch)
            frames_2, _, _ = collect_lif_frames(image, dapi_ch)

            slide = handle_slide_indexing(slice_start, slice_end, max_slice)
            
            # Output directory for results
            output_path = join(dirname(file_path), f"{base_name}_results")        
            makedirs(output_path, exist_ok=True)
            
            # Prepare target channel slice
            sample_slice_1 = np.max(frames_1[slide, :], axis=0)
            synaptotag_file_path = join(output_path, f"{base_name}_{im_index}_synaptotag.png")
            save_image(sample_slice_1, synaptotag_file_path, Step = "Target data", priority_keys=priority_keys)
            
            combined_image = np.zeros((*sample_slice_1.shape, 3), dtype='uint8')            
            sample_slice_1_normalized = (sample_slice_1 - np.min(sample_slice_1)) / (np.max(sample_slice_1) - np.min(sample_slice_1)) * 255
            combined_image[:, :, 0] = sample_slice_1_normalized.astype('uint8')  # RED: target_ch (synaptotagmin)

            # Check if DAPI channel is valid and add it if applicable
            if dapi_ch < max_channels:
                sample_slice_3 = np.max(frames_2[slide, :], axis=0)
                sample_slice_3_normalized = (sample_slice_3 - np.min(sample_slice_3)) / (np.max(sample_slice_3) - np.min(sample_slice_3)) * 255
                combined_image[:, :, 2] = sample_slice_3_normalized.astype('uint8')  # BLUE: dapi_ch (cell-label)
            
            combined_image_s.append(combined_image)

        return combined_image_s


def extract_image_stack(file_path, slice_start, slice_end, target_ch, dapi_ch):
    base_name = splitext(basename(file_path))[0]
    experiment_date = basename(dirname(file_path))    
    file_extension = splitext(file_path)[1].lower()    
    #print(f"target_ch: {target_ch}")
    #print(f"dapi_ch: {dapi_ch}")
    
    if file_extension == '.czi':
        
        combined_image_s = extract_czi_image_stack(file_path, slice_start, slice_end, target_ch, dapi_ch)
        
    elif file_extension == '.lif':        
        combined_image_s = extract_lif_stack(file_path, slice_start, slice_end, target_ch, dapi_ch)

    elif file_extension == '.tif':
        base_name = splitext(basename(file_path))[0]
        experiment_date = basename(dirname(file_path))    
        file_extension = splitext(file_path)[1].lower() 
        
        
        im_index = 0
        combined_image_s = []
        img = Image.open(file_path)        
        img_stack = [] 
        
        max_slice = img.n_frames
        slide = handle_slide_indexing(slice_start, slice_end, max_slice)       
        # Проходим по каждому слою в многослойном изображении
        for i in slide:
            img.seek(i)
            
            # Преобразуем слой в массив и нормализуем в 8-битный диапазон
            layer = np.array(img, dtype=np.uint16)
            layer_max = layer.max()
            
            # Проверка на максимальное значение, чтобы избежать деления на ноль
            if layer_max > 0:
                layer_normalized = (layer * (255.0 / layer_max)).astype(np.uint8)
            else:
                layer_normalized = layer.astype(np.uint8)  # Если max=0, просто переводим в uint8
                
            # Дублируем слой по трем каналам для RGB
            rgb_layer = np.stack([layer_normalized] * 3, axis=-1)
            img_stack.append(rgb_layer)
        
        # Объединяем все слои в одно RGB-изображение, используя максимумы по каждому каналу
        combined_image = np.max(np.stack(img_stack), axis=0)
        
        # Output directory for results
        output_path = join(dirname(file_path), f"{base_name}_results")        
        makedirs(output_path, exist_ok=True)
        
        # Save target channel slice
        synaptotag_file_path = join(output_path, f"{base_name}_{im_index}_synaptotag.png")
        save_image(combined_image, synaptotag_file_path, Step = "Target data", priority_keys=priority_keys)
            
        combined_image_s.append(combined_image)
        
    return combined_image_s

def load_coordinates_from_excel(excel_path, root):
    """
    Function to load coordinates from an Excel file.

    Arguments:
    excel_path -- path to the Excel file
    root -- the root window for the messagebox

    Returns:
    coords_df -- DataFrame with coordinates or None if there was an error
    """
    # Initialize coords_df as None
    coords_df = None

    # Check if the Excel file exists
    if os.path.exists(excel_path):
        # Try to load coordinates from the Excel file
        try:
            coords_df = pd.read_excel(excel_path, sheet_name='ROI_Coordinates')
            # Rename columns with duplicates
            #coords_df.columns = rename_column_names(coords_df.columns)        
        except Exception as e:
            # Show error message if the file could not be read
            messagebox.showerror("Error", f"Could not read coordinates from Excel file: {e}", parent=root)
            return None  # Stop execution if the file could not be read

    return coords_df

# Первый шаг
def stack_image(file_path, slice_start, slice_end, target_ch, dapi_ch):
    base_name = splitext(basename(file_path))[0]
    experiment_date = basename(dirname(file_path))
    combined_image_s = extract_image_stack(file_path, slice_start, slice_end, target_ch, dapi_ch)

    for im_index, combined_image in enumerate(combined_image_s):
        image_file_path = join(dirname(file_path), f"{base_name}_results", f"{base_name}_{im_index}_stack.png")
        # save RGB image
        combined_image = cv2.cvtColor(combined_image, cv2.COLOR_BGR2RGB)
        save_image(combined_image, image_file_path, Step = "Stack", priority_keys=stack_priority_keys)

# Загрузка параметров из временного файла
def load_params():
    if os.path.exists(TEMP_FILE):
        try:
            with open(TEMP_FILE, 'r') as f:
                return json.load(f)
        except (json.JSONDecodeError, KeyError):
            os.remove(TEMP_FILE)
    return {}

# Функция загрузки параметров из временного файла
def load_metadata():
    # Загружаем полный набор параметров
    params = load_params()
    # Ключи, которые нам нужны в metadata
    keys_to_extract = ['protocol', 'target_ch', 'selected_location', 'second_ch', 
                       'slice_start', 'slice_end', 'filter_radius', 'pixel_to_micron_ratio']
    # Создаем новый словарь с нужными ключами в порядке, в котором они указаны
    metadata = {key: params.get(key) for key in keys_to_extract}

    return metadata

# Накапливаем значения параметра и обновляем во временный файл
def collect_params(params, key, value):
    if key not in params:
        params[key] = []
    if value not in params[key]:  # Сохраняем только уникальные значения
        params[key].append(value)
    with open(TEMP_FILE, 'w') as f:
        json.dump(params, f)

# Сохранение параметров во временный файл
def save_params(params_dict):
    # Проверка, что передан именно словарь
    if not isinstance(params_dict, dict):
        raise ValueError("The argument passed must be a dictionary.")
    
    params = load_params()
    
    # Обновляем параметры новым словарем
    params.update(params_dict)
    
    with open(TEMP_FILE, 'w') as f:
        json.dump(params, f)

# Функция для записи данных в файл через ADS
def write_data_to_file(file_path, data):
    # Укажите путь к потоку данных как отдельную команду
    ads_path = f"{file_path}:syn_catch_metadata"
    # Используйте системные команды
    os.system(f'echo {json.dumps(data)} > "{ads_path}"')

def filter_files_by_metadata(folder_path, key, value):
    matching_files = []
    for root, dirs, files in os.walk(folder_path):
        for filename in files:
            file_path = os.path.join(root, filename)
            ads_path = f"{file_path}:syn_catch_metadata"
            if os.path.exists(ads_path):
                try:
                    with open(ads_path, 'r', encoding='utf-8') as f:
                        metadata = json.load(f)
                        if key in metadata and metadata[key] == value:
                            matching_files.append(file_path)
                except (json.JSONDecodeError, UnicodeDecodeError):
                    # Игнорируем файлы с некорректными метаданными
                    pass
    return matching_files

def get_location_name(root, location_names):
    """
    Функция отображает диалог для выбора или ввода имени региона.
    """
    # Рисуем окно
    dialog = initialize_window(root, "Select region", 350, 200, icon_path=icon_path)
    # Применяем тему через ThemeManager
    theme_manager = ThemeManager()
    theme_manager.apply_theme(dialog)
    dialog.attributes("-topmost", True)

    selected_location = [None]  # Для хранения выбранного значения

    def on_select():
        selected_location[0] = entry.get()
        dialog.destroy()  # Закрыть окно после выбора

    def on_option_select(option):
        entry.delete(0, "end")  # Очистить текущее значение в поле ввода
        entry.insert(0, option)  # Установить выбранное значение из списка

    # Метка
    label = Label(dialog, text="Select a region:")
    label.pack(pady=10)

    # Выпадающий список (OptionMenu)        
    if location_names:
        # Дублируем первый элемент массива location_names в начало
        location_names = [location_names[0]] + location_names
        selected_option = StringVar(dialog)
        selected_option.set('')  # Устанавливаем пустой элемент как начальное значение
        option_menu = OptionMenu(dialog, selected_option, *location_names, command=on_option_select)
        option_menu.pack(pady=5)

    # Поле для ввода текста
    entry = Entry(dialog)
    entry.insert(0, "region")  # Устанавливаем начальное значение "region"
    entry.pack(pady=5)

    # Кнопка подтверждения
    confirm_button = Button(dialog, text="OK", command=on_select)
    confirm_button.pack(pady=10)

    # Ожидание закрытия окна
    dialog.wait_window(dialog)

    return selected_location[0]



location_names = []

def select_location(file_path, root, initial_location = ''):
    scale_factor=0.8
           
    n_of_images = filetype_checking(file_path)  # Check number of images
    base_name = splitext(basename(file_path))[0]

    for im_index in range(n_of_images):
        excel_path = join(dirname(file_path), f"{base_name}_results", f"{base_name}_{im_index}_locations.xlsx")

        image_file_path = join(dirname(file_path), f"{base_name}_results", f"{base_name}_{im_index}_stack.png")
        image_in = read_image(image_file_path, as_pil = True, priority_keys=priority_keys).convert('RGB')
        image_np = np.array(image_in)

        selected_location = get_location_name(root, location_names)
        if not selected_location:
            print('Marking stopped')
            return pd.DataFrame(), initial_location  # Return an empty DataFrame
        # Сохраняем уникальное имя локации в локальных параметрах
        location_names.append(selected_location)
            
        coords_df = load_coordinates_from_excel(excel_path, root)
        
        # Initialize the drawing tool
        drawer = PolygonDrawer(image_np, scale_factor=scale_factor, coords_df=coords_df, comments = f"{base_name} #{im_index}")
        coords = drawer.run()  # Wait for drawing to complete

        if not coords:  # If the coordinates are empty
            print("No coordinates were drawn")
            return pd.DataFrame(), initial_location  # Return an empty DataFrame
        
        # сохраняем последний регион в глобальных параметрах
        save_params({'selected_location': selected_location})
        # Преобразуем координаты в DataFrame
        coords_df_new = pd.DataFrame(coords, columns=[f'{selected_location}_x', f'{selected_location}_y'])

        # Инициализируем ColorCycler
        color_cycler = ColorCycler(num_colors=10)
        #rgb_image = cv2.resize(image_np, None, fx=scale_factor, fy=scale_factor, interpolation=cv2.INTER_LINEAR)
        #bgr_image = cv2.cvtColor(self.rgb_image, cv2.COLOR_RGB2BGR)
        all_roi_img = image_np.copy()
        all_roi_img = cv2.cvtColor(all_roi_img, cv2.COLOR_RGB2BGR)
        all_roi_img = draw_polygons_on_image(coords_df, 1, color_cycler, all_roi_img, simplify_contour)
        all_roi_img = draw_polygons_on_image(coords_df_new, 1, color_cycler, all_roi_img, simplify_contour)
        all_roi_image_path = join(dirname(file_path), f"{base_name}_results", f"{base_name}_{im_index}_with_roi.png")
        save_image(all_roi_img, all_roi_image_path, Step = "Locations", priority_keys=stack_priority_keys)
        
        # Check if the Excel file exists to add new data
        if os.path.exists(excel_path):
            # Add new coordinates to the existing Excel file
            with pd.ExcelWriter(excel_path, mode='a', if_sheet_exists='overlay', engine='openpyxl') as writer:
                coords_df_new.to_excel(writer, sheet_name='ROI_Coordinates', startcol=writer.sheets['ROI_Coordinates'].max_column, index=False)
        else:
            # Create a new Excel file
            with pd.ExcelWriter(excel_path, engine='openpyxl') as writer:
                coords_df_new.to_excel(writer, sheet_name='ROI_Coordinates', index=False)

    print(f"Coordinates for {selected_location} successfully saved.")
    return coords_df_new, selected_location

# Второй шаг - фильтрация
def filter_after_roi_selection(filter_radius, file_path, location=''):
    
    base_name = splitext(basename(file_path))[0]
    experiment_date = basename(dirname(file_path))
       
    denoised_image_path_s = []
    n_of_images = filetype_checking(file_path)
    for im_index in range(n_of_images):
        # print(im_index)
        synaptotag_file_path = join(dirname(file_path), f"{base_name}_results", f"{base_name}_{im_index}_synaptotag.png")
        sample_slice_1 = read_image(synaptotag_file_path, priority_keys=priority_keys)
            
        # Convert image to uint8 format
        if sample_slice_1.dtype != np.uint8:
            sample_slice_1 = (255 * (sample_slice_1 - np.min(sample_slice_1)) / (np.max(sample_slice_1) - np.min(sample_slice_1))).astype(np.uint8)
        
        # Let's make sure the image is two-dimensional
        if len(sample_slice_1.shape) > 2:
            sample_slice_1 = cv2.cvtColor(sample_slice_1, cv2.COLOR_BGR2GRAY)
            
        # Remove background
        filter_radius = filter_radius | 1 # make odd
        background = cv2.medianBlur(sample_slice_1, filter_radius)
        denoised_image = cv2.subtract(sample_slice_1, background)
                
        denoised_image_path = join(dirname(file_path), f"{base_name}_results", f"{base_name}_{im_index}_denoised.png")
        
        save_image(denoised_image, denoised_image_path, Step = "Filtration",priority_keys=denoised_priority_keys)
        
        denoised_image_path_s.append(denoised_image_path)       
        
    return denoised_image_path_s

# Функция для переименования колонок, которые содержат суффиксы вида .1, .2
def rename_column_names(columns):
    new_columns = []
    
    for col in columns:
        if '.' in col:  # Если есть точка и номер
            # Разделяем название колонки на части
            base_name, suffix = col.split('.')
            # Переносим номер (suffix) перед '_x' или '_y'
            if '_x' in base_name:
                new_col = base_name.replace('_x', f'_{suffix}_x')
            elif '_y' in base_name:
                new_col = base_name.replace('_y', f'_{suffix}_y')
            else:
                new_col = base_name  # На случай, если нет "_x" или "_y"
        else:
            new_col = col  # Оставляем без изменений, если нет точки
        
        new_columns.append(new_col)
    
    return new_columns

def max_entropy_threshold(image):
    hist, bin_edges = histogram(image.ravel(), bins=256, density=True)
    cdf = hist.cumsum()
    cdf = cdf / cdf[-1]
    
    bin_mids = (bin_edges[:-1] + bin_edges[1:]) / 2.
    entropy = -hist * log(hist + finfo(float).eps)
    threshold = bin_mids[argmax(entropy)]
    return threshold

def get_threshold_value(image_array, binarization_method):
    if binarization_method == 'max_entropy':
        threshold_value = max_entropy_threshold(image_array)
    elif binarization_method == 'otsu':
        threshold_value = threshold_otsu(image_array)
    elif binarization_method == 'yen':
        threshold_value = threshold_yen(image_array)
    elif binarization_method == 'li':
        threshold_value = threshold_li(image_array)
    elif binarization_method == 'isodata':
        threshold_value = threshold_isodata(image_array)
    elif binarization_method == 'mean':
        threshold_value = threshold_mean(image_array)
    elif binarization_method == 'minimum':
        threshold_value = threshold_minimum(image_array)
    else:
        raise ValueError(f"Unsupported binarization method: {binarization_method}")
    
    return threshold_value

# Третий шаг - бинаризация
def binarize_images(file_path, binarization_method='max_entropy', min_size=64, max_size=100, pixel_to_micron_ratio = 0.12):    
    base_name = splitext(basename(file_path))[0]
    # distances = extract_scaling_distances_from_czi(file_path)
    # pixel_to_micron_ratio = distances['X']*1_000_000
    masks_image_path_s = []
    n_of_images = filetype_checking(file_path)
    for im_index in range(n_of_images):
        
        denoised_image_path = join(dirname(file_path), f"{base_name}_results", f"{base_name}_{im_index}_denoised.png")
        roi_coords_path = join(dirname(file_path), f"{base_name}_results", f"{base_name}_{im_index}_roi_coords.csv")
        
        image = read_image(denoised_image_path, as_pil = True, priority_keys=denoised_priority_keys).convert('L')
        image_array = array(image)        
        
        # get binary image
        threshold_value = get_threshold_value(image_array, binarization_method)    
        binary_image = image_array > threshold_value
        binary_image = remove_small_objects(binary_image, min_size=min_size)
        binary_image = remove_large_objects(binary_image, max_size=max_size)
        
        # Save full binary image
        full_binary_image_pil = Image.fromarray((binary_image * 255).astype('uint8'))
        full_binary_image_pil = ImageOps.invert(full_binary_image_pil)
        full_masks_image_path = join(dirname(file_path), f"{base_name}_results", f"{base_name}_{im_index}_full_masks_roi_crop.png")        
        save_image(full_binary_image_pil, full_masks_image_path, Step = "Full Binarization", priority_keys=full_binary_priority_keys)   
        
        coords_path = join(dirname(file_path), f"{base_name}_results", f"{base_name}_{im_index}_locations.xlsx")
        if os.path.exists(coords_path):
            # If coordinates file is found, load it
            coords_df = pd.read_excel(coords_path)
        else:          
            print('--------------------')
            print("WARNING:")
            print(f"No coordinates file found for image {im_index}. The entire area is taken for analysis")
            print('--------------------')
            return
        
        # Rename columns with duplicates
        coords_df.columns = rename_column_names(coords_df.columns)

        if coords_df is not None:
            for col_x in coords_df.columns[::2]:  # Iterate over pairs of x, y columns
                col_y = col_x.replace('_x', '_y')  # Find the matching y column
                location_name = col_x.rsplit('_', 1)[0]  # Extract the location name
                
                # Extract coordinates for the current location
                coords = coords_df[[col_x, col_y]].values
                
                # Create mask for the location
                roi_mask = create_region_mask(image_array.shape, coords)
                
                masks_image_path = join(dirname(file_path), f"{base_name}_results", f"{base_name}_{im_index}_{location_name}_masks_roi_crop.png")
                roi_mask_image_path = join(dirname(file_path), f"{base_name}_results", f"{base_name}_{im_index}_{location_name}_roi_mask.png")
                result_path = join(dirname(file_path), f"{base_name}_results", f"{base_name}_{im_index}_result_table.xlsx")
                
                save_params({'selected_location': location_name})
                
                # Save ROI mask as PNG image
                roi_mask_pil = Image.fromarray((roi_mask * 255).astype('uint8'))
                #save_image(roi_mask_pil, roi_mask_image_path, priority_keys=roi_mask_priority_keys)
                
                binary_image_roi = binary_image & roi_mask            
                binary_image_pil = Image.fromarray((binary_image_roi * 255).astype('uint8'))
                binary_image_pil = ImageOps.invert(binary_image_pil)                
                save_image(binary_image_pil, masks_image_path, Step = "Binarization", priority_keys=binary_image_priority_keys)               
                
                process_properties(location_name, 
                                   image_array, 
                                   binary_image_roi, 
                                   roi_mask, 
                                   pixel_to_micron_ratio, 
                                   binarization_method, 
                                   masks_image_path, 
                                   result_path)
                
                masks_image_path_s.append(masks_image_path)
        
    return masks_image_path_s


from openpyxl import load_workbook

def process_properties(location_name, 
                       image_array, 
                       binary_image_roi, 
                       roi_mask, 
                       pixel_to_micron_ratio, 
                       binarization_method, 
                       masks_image_path, 
                       result_path):
    metadata = load_metadata()
    labeled_image = label(binary_image_roi)
    props = regionprops(labeled_image, intensity_image=image_array)
    max_size = 500
    results = []
    total_objects = 0
    total_area = 0
    total_mean_intensity = 0
    roi_area = roi_mask.sum() * pixel_to_micron_ratio**2
    
    for index, prop in enumerate(props, start=1):
        area_microns = prop.area * pixel_to_micron_ratio**2
        if area_microns > max_size:
            continue
        total_objects += 1
        total_area += area_microns
        total_mean_intensity += prop.mean_intensity * area_microns
        
        result_dict = {
            "": index,
            "Area": f"{area_microns:.3f}",
            "Mean": f"{prop.mean_intensity:.3f}",
            "Min": int(prop.min_intensity),
            "Max": int(prop.max_intensity)
        }
        result_dict.update(metadata)
        results.append(result_dict)

    results_df = DataFrame(results)

    if total_objects > 0:
        average_size = total_area / total_objects
        average_mean_intensity = total_mean_intensity / total_area
    else:
        average_size = 0
        average_mean_intensity = 0

    summary_result = {
        "Slice": basename(masks_image_path),
        "Count": total_objects,
        "Total Area": f"{total_area:.3f}",
        "Average Size": f"{average_size:.3f}",
        "%Area": f"{(total_area / roi_area) * 100:.3f}",
        "Mean": f"{average_mean_intensity:.3f}",
        "Binarization method": binarization_method
    }

    hashes = generate_hashes_from_metadata(metadata, priority_keys)
    summary_result.update(metadata)
    new_summary_df = DataFrame([summary_result])

    if not os.path.exists(result_path):
        with pd.ExcelWriter(result_path, engine='openpyxl') as writer:
            new_summary_df.to_excel(writer, sheet_name="Summary", index=False)
            results_df.to_excel(writer, sheet_name=location_name, index=False)
    else:
        wb = load_workbook(result_path)
        all_sheets = {}

        for sh in wb.sheetnames:
            old_df = pd.read_excel(result_path, sheet_name=sh)
            all_sheets[sh] = old_df

        if "Summary" in all_sheets:
            all_sheets["Summary"] = pd.concat([all_sheets["Summary"], new_summary_df], ignore_index=True)
        else:
            all_sheets["Summary"] = new_summary_df

        if location_name in all_sheets:
            all_sheets[location_name] = pd.concat([all_sheets[location_name], results_df], ignore_index=True)
        else:
            all_sheets[location_name] = results_df

        with pd.ExcelWriter(result_path, engine='openpyxl', mode='w') as writer:
            if "Summary" in all_sheets:
                all_sheets["Summary"].to_excel(writer, sheet_name="Summary", index=False)
                del all_sheets["Summary"]
            if location_name in all_sheets:
                all_sheets[location_name].to_excel(writer, sheet_name=location_name, index=False)
                del all_sheets[location_name]
            for sheet_name, df_ in all_sheets.items():
                df_.to_excel(writer, sheet_name=sheet_name, index=False)



  
def gather_summary_files(file_path):
    base_name = splitext(basename(file_path))[0]
    results_folder = join(dirname(file_path), f"{base_name}_results")
    
    try:
        # Проверяем, существует ли папка с результатами
        summary_files = [f for f in os.listdir(results_folder) if '_summary_roi_result_table' in f and f.endswith('.xlsx')]
    except FileNotFoundError:
        # Сообщаем, что папка не найдена, и продолжаем выполнение
        print(f"Results folder not found for {file_path}. Skipping.")
        return []
    
    summary_data_s = []
    if summary_files:
        for summary_file in summary_files:
            summary_file_path = join(results_folder, summary_file)
            try:
                summary_data = pd.read_excel(summary_file_path)
                # Добавляем колонку с именем файла-источника
                summary_data['source_file'] = summary_file
                summary_data_s.append(summary_data)
            except Exception as e:
                print(f"Error reading file {summary_file_path}: {e}")
                continue
    else:
        # Сообщаем, что файлы не найдены, но продолжаем выполнение
        print(f"No summary data found for {file_path}. Skipping.")
        
    # Возвращаем собранные данные (или пустой список, если ничего не было собрано)
    return summary_data_s

def remove_large_objects(ar, max_size):
    # Label connected components
    labeled = label(ar)
    for region in regionprops(labeled):
        if region.area > max_size:
            ar[labeled == region.label] = 0
    return ar

# Function for single file post-processing
def pp_one(file_path, row, output_directory):
    
    experiment_date = basename(dirname(file_path))
    
    # Собираем данные из файлов с результатами
    summary_data_s = gather_summary_files(file_path)
    
    # Если данные есть, добавляем идентификаторы
    if summary_data_s:
        combined_summary_data = pd.concat(summary_data_s, ignore_index=True)
        combined_summary_data['ID'] = row['ID']
        combined_summary_data['Group'] = row['Group']
        

        return [combined_summary_data]
    
    # Если данных нет, возвращаем пустой список
    return []

def transform_parallelogram_to_rectangle(image, parallelogram_points, coords_df=None):
    # Convert parallelogram points to float32 for transformation
    parallelogram_points = np.array(parallelogram_points, dtype="float32")
    
    # Calculate the width and height of the parallelogram
    width = np.linalg.norm(parallelogram_points[1] - parallelogram_points[0])
    height = np.linalg.norm(parallelogram_points[2] - parallelogram_points[1])
    
    # Set the size of the resulting rectangle
    rectangle_size = (int(width), int(height))
    
    # Define the rectangle points in destination space
    rectangle_points = np.array([
        [0, 0],
        [rectangle_size[0] - 1, 0],
        [rectangle_size[0] - 1, rectangle_size[1] - 1],
        [0, rectangle_size[1] - 1]
    ], dtype="float32")
    
    # Get the transformation matrix from parallelogram to rectangle
    matrix = cv2.getPerspectiveTransform(parallelogram_points, rectangle_points)
    
    # Apply the perspective transformation to the image
    transformed_image = cv2.warpPerspective(image, matrix, rectangle_size)

    transformed_masks = {}

    # If coords_df is provided, create and transform masks for each location
    if coords_df is not None:
        for col_x in coords_df.columns[::2]:  # Iterate over pairs of x, y columns
            col_y = col_x.replace('_x', '_y')  # Find the matching y column
            location_name = col_x.rsplit('_', 1)[0]  # Extract the location name
            
            # Extract coordinates for the current location
            coords = coords_df[[col_x, col_y]].values
            
            # Create mask for the location
            mask = create_region_mask(image.shape, coords)
            
            # Transform the mask with the same transformation matrix
            transformed_mask = cv2.warpPerspective(mask, matrix, rectangle_size, flags=cv2.INTER_NEAREST)
            
            # Store the transformed mask in the dictionary with the location name
            transformed_masks[location_name] = transformed_mask

    return transformed_image, rectangle_size, transformed_masks

def calculate_histogram_for_color(image, axis, color_channel, bin_size=10):
    sum_brightness = np.sum(image[:, :, color_channel], axis=axis)
    count = np.ones_like(image[:, :, color_channel]).sum(axis=axis)
    mean_brightness = sum_brightness / count
    normalized_brightness = mean_brightness / 255
    binned_hist = np.add.reduceat(normalized_brightness, np.arange(0, normalized_brightness.shape[0], bin_size))
    return binned_hist

def calculate_histogram_for_gray(image, axis, bin_size=10, invert=False, mask=None):
    if invert:
        image = invert_image(image)

    # Convert image to grayscale by averaging across the color channels
    gray_image = np.mean(image, axis=2)
    
    if mask is not None:
        # Apply mask: keep only pixels where the mask is 1
        gray_image = gray_image * mask

    # Calculate the sum of brightness along the specified axis
    sum_brightness = np.sum(gray_image, axis=axis)

    # Count only non-zero (masked) pixels for normalization
    if mask is not None:
        count = np.sum(mask, axis=axis)  # Only count pixels where the mask is 1
    else:
        count = np.ones_like(gray_image).sum(axis=axis)  # Count all pixels if no mask

    # Prevent division by zero by replacing zero counts with ones (for regions fully masked out)
    count = np.maximum(count, 1)

    # Calculate mean brightness per row/column and normalize to [0, 1]
    mean_brightness = sum_brightness / count
    normalized_brightness = mean_brightness / 255

    # Apply binning to the brightness values
    binned_hist = np.add.reduceat(normalized_brightness, np.arange(0, normalized_brightness.shape[0], bin_size))
    
    return binned_hist

def plot_gray_histograms(image, rectangle_size, bin_size=10, invert=False, transformed_masks=None):    
    width, height = rectangle_size
    
    # Initialize a dictionary to store histogram data
    histograms_data = {}
    colors_used = {}  # Словарь для хранения цветов, использованных для каждой локации

    # Normalize the image proportions
    norm_width = width / max(width, height)
    norm_height = height / max(width, height)
    
    # Setting up the grid taking into account normalized image proportions
    fig = plt.figure(figsize=(8, 8))
    gs = gridspec.GridSpec(2, 2, width_ratios=[1, 1], height_ratios=[1, 1])
    
    legends = []
    cmap = plt.cm.get_cmap('tab10')  # Используем стандартную цветовую карту

    # Если нет масок, строим маску для всей области
    if transformed_masks is None or len(transformed_masks) == 0:
        transformed_masks = {'selected area': np.ones(image.shape[:2], dtype=np.uint8)}  # Маска для всей области

    # Цикл по каждой локации и её маске для построения гистограмм
    axes = ['x', 'y']  # Оси для построения гистограмм
    subplot_mapping = {'x': (gs[1, 1], 'X', width), 'y': (gs[0, 0], 'Y', height)}
    hist_axes = {}

    for axis in axes:
        ax = fig.add_subplot(subplot_mapping[axis][0])
        hist_axes[axis] = ax

        for idx, (location_name, mask) in enumerate(transformed_masks.items()):
            # Проверяем, есть ли ненулевые элементы в маске
            if np.any(mask):
                legends.append(location_name)

                # Расчет гистограммы для конкретной оси и локации
                hist_region = calculate_histogram_for_gray(image, axis=(0 if axis == 'x' else 1), 
                                                           bin_size=bin_size, invert=invert, mask=mask)

                # Построение гистограммы
                color = cmap(idx % 10)  # Получаем цвет из цветовой карты
                if axis == 'x':
                    ax.bar(np.arange(0, len(hist_region) * bin_size, bin_size), hist_region, 
                           width=bin_size, alpha=0.5, color=color)
                else:
                    ax.barh(np.arange(0, len(hist_region) * bin_size, bin_size), hist_region, 
                            height=bin_size, alpha=0.5, color=color)

                # Сохраняем данные гистограммы и цвет для локации
                if location_name not in histograms_data:
                    histograms_data[location_name] = {'hist_x': None, 'hist_y': None}
                    colors_used[location_name] = color  # Сохраняем цвет для локации
                histograms_data[location_name][f'hist_{axis}'] = hist_region

        # Установка заголовков для осей
        ax.set_title(f'Histogram along {subplot_mapping[axis][1]}-axis')
        if axis == 'x':
            ax.set_xlabel('Pixel position along X-axis')
            ax.set_ylabel('Normalized Mean Brightness')
            ax.set_xlim([0, width])
        else:
            ax.set_ylabel('Pixel position along Y-axis')
            ax.set_xlabel('Normalized Mean Brightness')
            ax.set_ylim([height, 0])  # Переворот оси Y

    # Отображение изображения с наложением локаций
    ax_image = fig.add_subplot(gs[0, 1])
    ax_image.imshow(cv2.cvtColor(image, cv2.COLOR_BGR2RGB), aspect='auto')
    ax_image.set_title("Transformed Rectangle Image")

    # Нанесение полигонов (локаций) на изображение с использованием накопленных цветов
    for location_name, mask in transformed_masks.items():
        if np.any(mask):
            color = to_rgba(colors_used[location_name], alpha=0.4)  # Используем сохраненный цвет для прозрачности
            contours, _ = cv2.findContours(mask, cv2.RETR_EXTERNAL, cv2.CHAIN_APPROX_SIMPLE)
            for contour in contours:
                # Преобразуем контуры в полигон для нанесения на изображение
                polygon = patches.Polygon(contour.reshape(-1, 2), closed=True, edgecolor=color, facecolor=color)
                ax_image.add_patch(polygon)

    # Добавление легенды
    if legends:
        hist_axes['x'].legend(legends, loc='upper right')
    
    plt.tight_layout()
    plt.show()
    
    return fig, histograms_data

def define_hist(file_path, location, slice_start, slice_end, target_ch, dapi_ch, root):
    base_name = splitext(basename(file_path))[0]
    experiment_date = basename(dirname(file_path))
    hist_image_path_s = []
    n_of_images = filetype_checking(file_path)

    combined_image_s = extract_image_stack(file_path, slice_start, slice_end, target_ch, dapi_ch)

    for im_index, combined_image in enumerate(combined_image_s):
        excel_path = join(dirname(file_path), f"{base_name}_results", f"{base_name}_{im_index}_locations.xlsx")
        if isfile(excel_path):
            coords_df = load_coordinates_from_excel(excel_path, root)
            if coords_df is not None:
                coords_df.columns = rename_column_names(coords_df.columns)
        else:
            coords_df = None

        editor = ParallelogramEditor(combined_image, scale_factor=0.8, coords_df=coords_df)

        # NEW: editor.run() сам вернет либо координаты, либо None, если была отмена
        parallelogram_points = editor.run()  
        if parallelogram_points is None:      # Если пользователь отменил
            continue                         # Пропускаем текущий im_index

        print("Read processed image")
        masks_image_path = join(dirname(file_path), f"{base_name}_results", f"{base_name}_{im_index}_full_masks_roi_crop.png")
        image = read_image(masks_image_path, priority_keys=full_binary_priority_keys)

        print("Transform processed image")
        transformed_image_rectangle, rectangle_size, transformed_masks = transform_parallelogram_to_rectangle(
            image, parallelogram_points, coords_df
        )

        print("Calculate histogram")
        fig, histograms_data = plot_gray_histograms(
            transformed_image_rectangle, rectangle_size, bin_size=10, invert=True, transformed_masks=transformed_masks
        )

        hist_image_path = join(dirname(file_path), f"{base_name}_results", f"{base_name}_{im_index}_hist.png")
        save_image(fig, hist_image_path, Step='Histogram', priority_keys=full_binary_priority_keys)
        hist_image_path_s.append(hist_image_path)

        hist_table_path = join(dirname(file_path), f"{base_name}_results", f"{base_name}_{im_index}_histograms.xlsx")
        with pd.ExcelWriter(hist_table_path, engine='openpyxl') as writer:
            for location_name, hist_data in histograms_data.items():
                hist_df = pd.DataFrame({
                    'Pixel Position (X)': pd.Series(hist_data['hist_x']),
                    'Pixel Position (Y)': pd.Series(hist_data['hist_y']),
                    'Parallelogram coords (X,Y)': pd.Series(parallelogram_points)
                })
                hist_df.to_excel(writer, sheet_name=location_name, index=False)

    return hist_image_path_s


TEMP_FILE = os.path.join(tempfile.gettempdir(), 'synapto_catch_params.json')
local_params = {'location_names': []}

# Ключи для установления иерархии сохранения
priority_keys = ['protocol', 'target_ch', 'slice_start', 'slice_end']
target_priority_keys = ['protocol', 'target_ch', 'slice_start', 'slice_end']
stack_priority_keys = priority_keys + ['second_ch']
denoised_priority_keys = priority_keys + ['filter_radius']
full_binary_priority_keys = priority_keys + ['filter_radius']
roi_mask_priority_keys = full_binary_priority_keys + ['selected_location']
binary_image_priority_keys = roi_mask_priority_keys + ['selected_location']